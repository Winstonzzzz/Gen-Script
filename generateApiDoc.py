import os
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border, Side
from openpyxl import Workbook

# 定义解析注解和类字段的正则表达式
api_model_property_pattern = re.compile(r'@ApiModelProperty\(value\s*=\s*"([^"]+)"\)')
json_property_pattern = re.compile(r'@JsonProperty\("([^"]+)"\)')
field_pattern = re.compile(r'private\s+([\w<>[\],\s]+)\s+([\w]+);')

# 定义Excel模版文件路径和输出目录
output_dir = os.path.expanduser('~/Documents/脚本/output/')

# 设置黄色填充
yellow_fill = PatternFill(start_color="FFFD54", end_color="FFFD54", fill_type="solid")
# 设置稍暗的绿色填充
green_fill = PatternFill(start_color="4EAC5B", end_color="4EAC5B", fill_type="solid")

# 解析Java文件函数
def parse_java_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    fields = []
    # 提取所有字段定义
    field_info = field_pattern.findall(content)
    # 提取所有ApiModelProperty注解值
    api_model_properties = api_model_property_pattern.findall(content)
    # 提取所有JsonProperty注解值
    json_properties = json_property_pattern.findall(content)
    for i in range(len(field_info)):
        field_type = field_info[i][0]
        field_name = field_info[i][1]
        description = api_model_properties[i]
        english_name = json_properties[i] if json_properties[i] else field_name

        fields.append((english_name, description, field_type))

    return fields

def init_excel():
    # 加载Excel模版
    wb = Workbook()
    ws = wb.active

    # 设置接口名称
    ws['A1'] = '接口名称'
    ws['A2'] = '接口中文名'

    ws['A4'] = '英文名称'
    ws['A4'].fill = yellow_fill
    ws['B4'] = '中文名称'
    ws['B4'].fill = yellow_fill
    ws['C4'] = '数据类型'
    ws['C4'].fill = yellow_fill
    ws['D4'] = '长度'
    ws['D4'].fill = yellow_fill
    ws['E4'] = '是否必输'
    ws['E4'].fill = yellow_fill
    ws['F4'] = '枚举值'
    ws['F4'].fill = yellow_fill
    ws['G4'] = '备注'
    ws['G4'].fill = yellow_fill

    ws['A5'] = '输入'
    ws['A5'].fill = green_fill
    ws['B5'].fill = green_fill
    ws['C5'].fill = green_fill
    ws['D5'].fill = green_fill
    ws['E5'].fill = green_fill
    ws['F5'].fill = green_fill
    ws['G5'].fill = green_fill

    return wb,ws

# 生成Excel文件函数
def generate_excel(wb, ws, interface_name, request_fields, response_fields):
    ws['B1'] = interface_name

    # 填充Request字段
    row_idx = 6
    for english_name, description, data_type in request_fields:
        ws[f'A{row_idx}'] = description
        ws[f'B{row_idx}'] = english_name
        ws[f'C{row_idx}'] = data_type
        row_idx += 1

    # 填充Response字段
    ws[f'A{row_idx}'] = '输出'
    ws[f'A{row_idx}'].fill = green_fill
    ws[f'B{row_idx}'].fill = green_fill
    ws[f'C{row_idx}'].fill = green_fill
    ws[f'D{row_idx}'].fill = green_fill
    ws[f'E{row_idx}'].fill = green_fill
    ws[f'F{row_idx}'].fill = green_fill
    ws[f'G{row_idx}'].fill = green_fill
    row_idx += 1

    for english_name, description, data_type in response_fields:
        ws[f'A{row_idx}'] = description
        ws[f'B{row_idx}'] = english_name
        ws[f'C{row_idx}'] = data_type
        row_idx += 1

    # 定义边框样式
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # 定义要加边框的区域
    start_row = 1
    start_column = 1
    end_row = row_idx - 1
    end_column = 7

    # 给指定区域的所有单元格加上边框线
    for row in ws.iter_rows(min_row=start_row, min_col=start_column, max_row=end_row, max_col=end_column):
        for cell in row:
            cell.border = border_style

    # 保存Excel文件
    output_path = os.path.join(output_dir, f'{interface_name}.xlsx')
    wb.save(output_path)

# 主函数
def main(request_dir, response_dir):
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 获取所有请求文件和响应文件
    request_files = [f for f in os.listdir(request_dir) if f.endswith('Request.java')]
    response_files = [f for f in os.listdir(response_dir) if f.endswith('Response.java')]

    for request_file in request_files:
        interface_name = request_file.replace('Request.java', '')
        response_file = f'{interface_name}Response.java'

        if response_file in response_files:
            request_fields = parse_java_file(os.path.join(request_dir, request_file))
            response_fields = parse_java_file(os.path.join(response_dir, response_file))

            wb,ws = init_excel()

            generate_excel(wb,ws, interface_name, request_fields, response_fields)
        else:
            print(f'No matching response file for {request_file}')

# 执行主函数
if __name__ == '__main__':
    print("请输入请求文件路径 (request_dir):")
    request_dir = input().strip()
    print("请输入响应文件路径 (response_dir):")
    response_dir = input().strip()
    main(request_dir, response_dir)
    # request_dir = '/Users/wshuobangmacpro/Documents/脚本/request'  # 替换为实际请求文件路径
    # response_dir = '/Users/wshuobangmacpro/Documents/脚本/response'  # 替换为实际响应文件路径
